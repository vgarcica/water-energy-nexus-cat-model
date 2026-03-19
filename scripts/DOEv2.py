# -*- coding: utf-8 -*-
"""
Created on Fri Jan 16 13:43:30 2026

@author: tirki
"""

import pandas as pd
import numpy as np
import itertools
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
import warnings
warnings.filterwarnings('ignore')

class DOEEnergetico:
    """
    Clase para Diseño de Experimentos en optimización energética
    """
    
    def __init__(self, procesar_escenario_func, datos_base):
        self.procesar_escenario = procesar_escenario_func
        self.datos_base = datos_base
        self.factores_definicion = {}
        self.matriz_experimentos = None
        self.resultados_experimentos = None
        self.efectos_principales = {}
        self.interacciones = {}
        
    def definir_factores(self):
        """
        Define los 7 factores con sus 2 niveles cada uno
        """
        # Obtener valores actuales del DataFrame de potencia
        potencia_actual = self.datos_base['df_potencia']
        fotovoltaica_actual = potencia_actual.Fotovoltaica.iloc[-1] + potencia_actual.Termosolar.iloc[-1]
        eolica_actual = potencia_actual.Eòlica.iloc[-1]
        
        self.factores_definicion = {
            'fotovoltaica': {
                'nombre': 'Fotovoltaica',
                'nivel_bajo': 500, #fotovoltaica_actual,  # Nivel actual
                'nivel_alto': 5000,
                'tipo': 'continuo'
            },
            'eolica': {
                'nombre': 'Eólica', 
                'nivel_bajo': 1500, #eolica_actual,  # Nivel actual
                'nivel_alto': 5000,
                'tipo': 'continuo'
            },
            'baterias': {
                'nombre': 'Baterías/Bombeos',
                'nivel_bajo': 500,  # Actual
                'nivel_alto': 2000,
                'tipo': 'continuo'
            },
            'desalacion': {
                'nombre': 'Desalación',
                'nivel_bajo': 30,
                'nivel_alto': 60,
                'tipo': 'continuo'
            },
            'min_run_hours': {
                'nombre': 'Min Run Hours',
                'nivel_bajo': 6,
                'nivel_alto': 12,
                'tipo': 'discreto'
            },
            'midpoint': {
                'nombre': 'Midpoint',
                'nivel_bajo': 60,
                'nivel_alto': 75,
                'tipo': 'continuo'
            },
            'overflow_threshold': {
                'nombre': 'Overflow Threshold',
                'nivel_bajo': 75,
                'nivel_alto': 95,
                'tipo': 'continuo'
            }
        }
        
        print("Factores definidos:")
        for key, factor in self.factores_definicion.items():
            print(f"  {factor['nombre']}: {factor['nivel_bajo']} → {factor['nivel_alto']}")
    
    # def generar_matriz_experimentos(self):
    #     """
    #     Genera la matriz de experimentos (2^7 = 128 experimentos)
    #     """
    #     factores = list(self.factores_definicion.keys())
    #     n_factores = len(factores)
        
    #     # Generar todas las combinaciones posibles (-1, +1)
    #     combinaciones = list(itertools.product([-1, 1], repeat=n_factores))
        
    #     # Crear DataFrame con la matriz de experimentos
    #     self.matriz_experimentos = pd.DataFrame(combinaciones, columns=factores)
        
    #     # Añadir columnas con valores reales
    #     for factor in factores:
    #         factor_def = self.factores_definicion[factor]
    #         nivel_bajo = factor_def['nivel_bajo']
    #         nivel_alto = factor_def['nivel_alto']
            
    #         # Convertir de codificación (-1, +1) a valores reales
    #         self.matriz_experimentos[f'{factor}_real'] = self.matriz_experimentos[factor].apply(
    #             lambda x: nivel_bajo if x == -1 else nivel_alto
    #         )
        
    #     # Añadir columna de experimento
    #     self.matriz_experimentos['experimento'] = range(1, len(self.matriz_experimentos) + 1)
        
    #     print(f"Matriz de experimentos generada: {len(self.matriz_experimentos)} experimentos")
    #     return self.matriz_experimentos
    
    def generar_matriz_experimentos(self):
        """
        Genera la matriz de experimentos en ordre de Yates (2^7 = 128 experiments).
        Inclou etiquetes estàndard de disseny factorial.
        """
        factores = list(self.factores_definicion.keys())
        n_factores = len(factores)
        letras = 'ABCDEFG'[:n_factores]
        
        # Generar combinacions en ORDRE DE YATES
        # L'ordre de Yates correspon a comptar en binari, 
        # on el bit menys significatiu és el primer factor
        combinaciones = []
        etiquetas = []
        
        for i in range(2**n_factores):
            combo = []
            etiqueta_parts = []
            
            for j in range(n_factores):
                if (i >> j) & 1:  # Bit j actiu
                    combo.append(1)
                    etiqueta_parts.append(letras[j].lower())
                else:
                    combo.append(-1)
            
            combinaciones.append(combo)
            
            # Etiqueta: (1) si tots a nivell baix, sinó concatenació de lletres
            if not etiqueta_parts:
                etiquetas.append('(1)')
            else:
                etiquetas.append(''.join(etiqueta_parts))
        
        # Crear DataFrame
        self.matriz_experimentos = pd.DataFrame(combinaciones, columns=factores)
        
        # Afegir etiqueta i número d'experiment
        self.matriz_experimentos.insert(0, 'etiqueta', etiquetas)
        self.matriz_experimentos.insert(0, 'experimento', range(1, len(combinaciones) + 1))
        
        # Afegir columnes amb valors reals
        for factor in factores:
            factor_def = self.factores_definicion[factor]
            nivel_bajo = factor_def['nivel_bajo']
            nivel_alto = factor_def['nivel_alto']
            
            self.matriz_experimentos[f'{factor}_real'] = self.matriz_experimentos[factor].apply(
                lambda x: nivel_bajo if x == -1 else nivel_alto
            )
        
        print(f"Matriu d'experiments generada: {len(self.matriz_experimentos)} experiments (ordre de Yates)")
        print(f"Primeres etiquetes: {etiquetas[:8]}")
        
        return self.matriz_experimentos    
    
    # def ejecutar_experimentos(self, experimentos_subset=None, verbose=True):
    #     """
    #     Ejecuta todos los experimentos y almacena resultados.
    #     """
    #     if self.matriz_experimentos is None:
    #         self.generar_matriz_experimentos()
        
    #     if experimentos_subset is None:
    #         experimentos_a_ejecutar = self.matriz_experimentos
    #     else:
    #         experimentos_a_ejecutar = self.matriz_experimentos.iloc[experimentos_subset]
        
    #     resultados = []
        
    #     print(f"Executant {len(experimentos_a_ejecutar)} experiments...")
        
    #     for idx, row in experimentos_a_ejecutar.iterrows():
    #         if verbose and (idx + 1) % 10 == 0:
    #             print(f"  Progrés: {idx + 1}/{len(experimentos_a_ejecutar)}")
            
    #         try:
    #             params = {
    #                 'potencia_solar': row['fotovoltaica_real'],
    #                 'potencia_eolica': row['eolica_real'], 
    #                 'potencia_baterias': row['baterias_real'],
    #                 'max_desalation': row['desalacion_real'],
    #                 'min_run_hours': int(row['min_run_hours_real']),
    #                 'midpoint_estimation': row['midpoint_real'],
    #                 'overflow_threshold_pct': row['overflow_threshold_real']
    #             }
                
    #             resultado = self.procesar_escenario(**self.datos_base, **params)
    #             objetivos = self.extraer_objetivos(resultado)
                
    #             fila_resultado = {
    #                 'experimento': row['experimento'],
    #                 **{f'{k}': row[k] for k in self.factores_definicion.keys()},
    #                 **{f'{k}_real': row[f'{k}_real'] for k in self.factores_definicion.keys()},
    #                 **objetivos,
    #                 'valido': True
    #             }
                
    #         except Exception as e:
    #             if verbose:
    #                 print(f"    Error en experiment {row['experimento']}: {e}")
                
    #             fila_resultado = {
    #                 'experimento': row['experimento'],
    #                 **{f'{k}': row[k] for k in self.factores_definicion.keys()},
    #                 **{f'{k}_real': row[f'{k}_real'] for k in self.factores_definicion.keys()},
    #                 'min_nivell': np.nan,
    #                 'termica_imports': np.nan,
    #                 'excedents': np.nan,
    #                 'costos': np.nan,
    #                 'factor_capacitat': np.nan,
    #                 'amplitud_estacional': np.nan,
    #                 'msd_nivell': np.nan,
    #                 'valido': False
    #             }
            
    #         resultados.append(fila_resultado)
        
    #     self.resultados_experimentos = pd.DataFrame(resultados)
    #     print(f"Experiments completats: {self.resultados_experimentos['valido'].sum()} exitosos")
    #     return self.resultados_experimentos

    
    def extraer_objetivos(self, resultado):
        """
        Extrae los 7 objetivos del resultado de la simulación.
        Retorna valores naturales (no transformados) para visualització.
        """
        
        # 1. Mínim nivell [%] (a maximitzar → guardarem valor natural)
        min_nivell = resultado['level_final'].min()
        
        # 2. Generació tèrmica + importacions [MWh] (a minimitzar)
        termica_imports = resultado['energy_data']['Gas+Imports'].sum()
        
        # 3. Excedents post-dessalinització [MWh] (a minimitzar)
        excedents = resultado.get('excedents', resultado.get('surpluses_net', 0))
        if hasattr(excedents, 'sum'):
            excedents_total = excedents.sum()
        else:
            excedents_total = excedents
        
        # 4. Costos d'instal·lació [€] (a minimitzar)
        costos = resultado['costes']['total']
        
        # 5. Factor de capacitat dessalinització [%] (a maximitzar → valor natural)
        factor_capacitat = resultado.get('capacity_factor', 0)
        
        # 6. Amplitud estacional [%] (a minimitzar)
        amplitud_estacional = resultado['hydro_metrics']['Variación estacional (%)']
        
        # 7. Mean Squared Deviation del nivell (a minimitzar)
        level_final = resultado['level_final']
        msd_nivell = ((100 - level_final)**2).sum() / len(level_final)
        
        return {
            'min_nivell': min_nivell,                    # Valor natural [%]
            'termica_imports': termica_imports,          # [MWh]
            'excedents': excedents_total,                # [MWh]
            'costos': costos,                            # [€]
            'factor_capacitat': factor_capacitat,        # Valor natural [%]
            'amplitud_estacional': amplitud_estacional,  # [%]
            'msd_nivell': msd_nivell,                    # [%²]
        }
    
    def calcular_efectos_principales(self):
        """
        Calcula los efectos principales de cada factor sobre cada objetivo.
        Transforma internament els objectius a minimitzar per coherència.
        """
        if self.resultados_experimentos is None:
            raise ValueError("Primero ejecuta los experimentos")
        
        datos_validos = self.resultados_experimentos[self.resultados_experimentos['valido']].copy()
        
        # Definir objectius i si cal invertir-los per l'anàlisi
        objetivos_config = {
            'min_nivell': {'invertir': True, 'desc': 'Mínim nivell (MAX)'},
            'termica_imports': {'invertir': False, 'desc': 'Tèrmica+Imports (MIN)'},
            'excedents': {'invertir': False, 'desc': 'Excedents (MIN)'},
            'costos': {'invertir': False, 'desc': 'Costos (MIN)'},
            'factor_capacitat': {'invertir': True, 'desc': 'Factor capacitat (MAX)'},
            'amplitud_estacional': {'invertir': False, 'desc': 'Amplitud estacional (MIN)'},
            'msd_nivell': {'invertir': False, 'desc': 'MSD nivell (MIN)'},
        }
        
        factores = list(self.factores_definicion.keys())
        efectos = {}
        
        for objetivo, config in objetivos_config.items():
            if objetivo not in datos_validos.columns or datos_validos[objetivo].isna().all():
                continue
            
            efectos_obj = {}
            
            for factor in factores:
                nivel_alto = datos_validos[datos_validos[factor] == 1][objetivo].mean()
                nivel_bajo = datos_validos[datos_validos[factor] == -1][objetivo].mean()
                
                efecto = nivel_alto - nivel_bajo
                
                # # Invertir signe si l'objectiu és a maximitzar (per coherència: efecte positiu = millora)
                # if config['invertir']:
                #     efecto = -efecto
                
                efectos_obj[factor] = efecto
            
            efectos[objetivo] = efectos_obj
        
        self.efectos_principales = efectos
        self.objetivos_config = objetivos_config
        return efectos
    
    def calcular_interacciones_2way(self):
        """
        Calcula las interacciones de 2 factores.
        """
        if self.resultados_experimentos is None:
            raise ValueError("Primero ejecuta los experimentos")
        
        datos_validos = self.resultados_experimentos[self.resultados_experimentos['valido']].copy()
        factores = list(self.factores_definicion.keys())
        
        # Obtenir llista d'objectius vàlids
        objetivos = ['min_nivell', 'termica_imports', 'excedents', 'costos', 
                     'factor_capacitat', 'amplitud_estacional', 'msd_nivell']
        
        interacciones = {}
        
        for objetivo in objetivos:
            if objetivo not in datos_validos.columns or datos_validos[objetivo].isna().all():
                continue
            
            interacciones_obj = {}
            
            for i, factor1 in enumerate(factores):
                for factor2 in factores[i+1:]:
                    y_pp = datos_validos[(datos_validos[factor1] == 1) & (datos_validos[factor2] == 1)][objetivo].mean()
                    y_pn = datos_validos[(datos_validos[factor1] == 1) & (datos_validos[factor2] == -1)][objetivo].mean()
                    y_np = datos_validos[(datos_validos[factor1] == -1) & (datos_validos[factor2] == 1)][objetivo].mean()
                    y_nn = datos_validos[(datos_validos[factor1] == -1) & (datos_validos[factor2] == -1)][objetivo].mean()
                    
                    interaccion = (y_pp - y_pn - y_np + y_nn) / 2 #Dividir entre 4 si volem obtenir el pendent
                    interacciones_obj[f'{factor1}×{factor2}'] = interaccion
            
            interacciones[objetivo] = interacciones_obj
        
        self.interacciones = interacciones
        return interacciones
    
    def generar_reporte_completo(self, top_n_efectos=5):
        """
        Genera un reporte completo amb efectes i interaccions.
        """
        if not self.efectos_principales:
            self.calcular_efectos_principales()
        if not self.interacciones:
            self.calcular_interacciones_2way()
        
        print("="*80)
        print("REPORTE DE DISEÑO DE EXPERIMENTOS")
        print("="*80)
        
        # Resumen
        total_exp = len(self.resultados_experimentos)
        validos = self.resultados_experimentos['valido'].sum()
        print(f"\nRESUM D'EXPERIMENTS:")
        print(f"  Total executats: {total_exp}")
        print(f"  Exitosos: {validos}")
        print(f"  Fallits: {total_exp - validos}")
        
        # Estadístiques dels objectius
        print(f"\n{'='*60}")
        print("ESTADÍSTIQUES DELS OBJECTIUS")
        print(f"{'='*60}")
        
        datos_validos = self.resultados_experimentos[self.resultados_experimentos['valido']]
        
        for objetivo, config in self.objetivos_config.items():
            if objetivo in datos_validos.columns:
                vals = datos_validos[objetivo]
                print(f"\n{config['desc']}:")
                print(f"  Rang: [{vals.min():.2f}, {vals.max():.2f}]")
                print(f"  Mitjana: {vals.mean():.2f} ± {vals.std():.2f}")
        
        # Efectes principals per objectiu
        for objetivo, efectos in self.efectos_principales.items():
            config = self.objetivos_config.get(objetivo, {'desc': objetivo})
            print(f"\n{'-'*60}")
            print(f"{config['desc']}")
            print(f"{'-'*60}")
            
            efectos_ordenados = sorted(efectos.items(), key=lambda x: abs(x[1]), reverse=True)
            
            print("EFECTES PRINCIPALS (Top 5):")
            for factor, efecto in efectos_ordenados[:top_n_efectos]:
                factor_nombre = self.factores_definicion[factor]['nombre']
                # Efecte positiu = millora l'objectiu (ja sigui augmentar o reduir segons el cas)
                direccion = "✓" if efecto > 0 else "✗"
                print(f"  {factor_nombre:20s}: {efecto:+12.2f} {direccion}")
            
            if objetivo in self.interacciones:
                interac_ordenadas = sorted(self.interacciones[objetivo].items(), 
                                         key=lambda x: abs(x[1]), reverse=True)
                
                print(f"\nINTERACCIONS 2-WAY (Top 3):")
                for interac, valor in interac_ordenadas[:3]:
                    print(f"  {interac:20s}: {valor:+12.2f}")
    
    # def exportar_resultados(self, nombre_archivo="resultados_doe.xlsx"):
    #     """
    #     Exporta todos los resultados a Excel
    #     """
    #     with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
    #         # Matriz de experimentos y resultados
    #         self.resultados_experimentos.to_excel(writer, sheet_name='Experimentos', index=False)
            
    #         # Efectos principales
    #         efectos_df = pd.DataFrame(self.efectos_principales).T
    #         efectos_df.to_excel(writer, sheet_name='Efectos_Principales')
            
    #         # Interacciones
    #         interacciones_df = pd.DataFrame(self.interacciones).T
    #         interacciones_df.to_excel(writer, sheet_name='Interacciones_2Way')
        
    #     print(f"Resultados exportados a: {nombre_archivo}")
        
    def ejecutar_experimentos(self, experimentos_subset=None, verbose=True):
        """
        Ejecuta todos los experimentos y almacena resultados.
        """
        if self.matriz_experimentos is None:
            self.generar_matriz_experimentos()
        
        if experimentos_subset is None:
            experimentos_a_ejecutar = self.matriz_experimentos
        else:
            experimentos_a_ejecutar = self.matriz_experimentos.iloc[experimentos_subset]
        
        resultados = []
        
        print(f"Executant {len(experimentos_a_ejecutar)} experiments...")
        
        for idx, row in experimentos_a_ejecutar.iterrows():
            if verbose and (idx + 1) % 10 == 0:
                print(f"  Progrés: {idx + 1}/{len(experimentos_a_ejecutar)}")
            
            try:
                params = {
                    'potencia_solar': row['fotovoltaica_real'],
                    'potencia_eolica': row['eolica_real'], 
                    'potencia_baterias': row['baterias_real'],
                    'max_desalation': row['desalacion_real'],
                    'min_run_hours': int(row['min_run_hours_real']),
                    'midpoint_estimation': row['midpoint_real'],
                    'overflow_threshold_pct': row['overflow_threshold_real']
                }
                
                resultado = self.procesar_escenario(**self.datos_base, **params)
                objetivos = self.extraer_objetivos(resultado)
                
                fila_resultado = {
                    'experimento': row['experimento'],
                    'etiqueta': row['etiqueta'],  # AFEGIT
                    **{f'{k}': row[k] for k in self.factores_definicion.keys()},
                    **{f'{k}_real': row[f'{k}_real'] for k in self.factores_definicion.keys()},
                    **objetivos,
                    'valido': True
                }
                
            except Exception as e:
                if verbose:
                    print(f"    Error en experiment {row['experimento']}: {e}")
                
                fila_resultado = {
                    'experimento': row['experimento'],
                    'etiqueta': row['etiqueta'],  # AFEGIT
                    **{f'{k}': row[k] for k in self.factores_definicion.keys()},
                    **{f'{k}_real': row[f'{k}_real'] for k in self.factores_definicion.keys()},
                    'min_nivell': np.nan,
                    'termica_imports': np.nan,
                    'excedents': np.nan,
                    'costos': np.nan,
                    'factor_capacitat': np.nan,
                    'amplitud_estacional': np.nan,
                    'msd_nivell': np.nan,
                    'valido': False
                }
            
            resultados.append(fila_resultado)
        
        self.resultados_experimentos = pd.DataFrame(resultados)
        print(f"Experiments completats: {self.resultados_experimentos['valido'].sum()} exitosos")
        return self.resultados_experimentos
    
    
    def exportar_resultados(self, nombre_archivo="resultados_doe.xlsx"):
        """
        Exporta todos los resultados a Excel amb etiquetes.
        """
        # Mapeig factors → lletres
        factores = list(self.factores_definicion.keys())
        letras = 'ABCDEFG'[:len(factores)]
        factor_a_letra = {f: letras[i] for i, f in enumerate(factores)}
        
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            # Pestanya 1: Experiments i resultats
            self.resultados_experimentos.to_excel(writer, sheet_name='Experiments', index=False)
            
            # Pestanya 2: Efectes principals
            efectos_df = pd.DataFrame(self.efectos_principales).T
            efectos_df.index.name = 'Objectiu'
            efectos_df.to_excel(writer, sheet_name='Efectes_Principals', startrow=1)
            
            # Afegir fila amb lletres (A, B, C, D, E, F, G)
            ws_efectes = writer.sheets['Efectes_Principals']
            for i, factor in enumerate(efectos_df.columns):
                letra = factor_a_letra.get(factor, factor)
                ws_efectes.cell(row=1, column=i+2, value=letra)  # +2 perquè columna 1 és l'índex
            
            # Pestanya 3: Interaccions 2-way
            interacciones_df = pd.DataFrame(self.interacciones).T
            interacciones_df.index.name = 'Objectiu'
            interacciones_df.to_excel(writer, sheet_name='Interaccions_2Way', startrow=1)
            
            # Afegir fila amb parelles de lletres (AB, AC, BC, etc.)
            ws_inter = writer.sheets['Interaccions_2Way']
            for i, interaccion in enumerate(interacciones_df.columns):
                # Convertir 'factor1×factor2' a 'AB'
                parts = interaccion.split('×')
                letra1 = factor_a_letra.get(parts[0], parts[0][0].upper())
                letra2 = factor_a_letra.get(parts[1], parts[1][0].upper())
                etiqueta = letra1 + letra2
                ws_inter.cell(row=1, column=i+2, value=etiqueta)
            
            # Pestanya 4: Resum experiments x objectius
            cols_objectius = ['min_nivell', 'termica_imports', 'excedents', 'costos', 
                             'factor_capacitat', 'amplitud_estacional', 'msd_nivell']
            resum_df = self.resultados_experimentos[['experimento', 'etiqueta'] + cols_objectius].copy()
            resum_df.to_excel(writer, sheet_name='Resum_Objectius', index=False)
        
        print(f"Resultats exportats a: {nombre_archivo}")

    # Función principal para ejecutar el DOE
    def ejecutar_doe_completo(procesar_escenario_func, datos_base, experimentos_prueba=None):
        """
        Función principal para ejecutar el diseño de experimentos completo
        
        Args:
            procesar_escenario_func: Tu función procesar_escenario
            datos_base: Diccionario con todos los datos base
            experimentos_prueba: Número de experimentos para prueba (None = todos los 128)
        """
        
        # Crear objeto DOE
        doe = DOEEnergetico(procesar_escenario_func, datos_base)
        
        # Definir factores
        doe.definir_factores()
        
        # Generar matriz de experimentos
        matriz = doe.generar_matriz_experimentos()
        
        # Ejecutar experimentos (subset para prueba si se especifica)
        if experimentos_prueba:
            subset_indices = np.random.choice(len(matriz), size=experimentos_prueba, replace=False)
            print(f"Ejecutando subset de {experimentos_prueba} experimentos para prueba...")
        else:
            subset_indices = None
            print("Ejecutando todos los 128 experimentos...")
        
        resultados = doe.ejecutar_experimentos(subset_indices)
        
        # Calcular efectos e interacciones
        doe.calcular_efectos_principales()
        doe.calcular_interacciones_2way()
        
        # Generar reporte
        doe.generar_reporte_completo()
        
        # Exportar resultados
        doe.exportar_resultados(nombre_archivo="resultats_DOE.xlsx")
        
        return doe

#%%

if __name__ == "__main__":
    
    datos_base = {
        # Sèries de dades
        'df_demanda': datos.demanda,
        'df_nuclear': datos.nuclears_base,
        'df_cogeneracion': datos.cogeneracion_h,
        'df_solar': datos.solar_h,
        'df_eolica': datos.eolica_h,
        'df_autoconsum': datos.autoconsum_hourly,
        'df_potencia': datos.potencia,
        'df_niveles_int': datos.df_pct_int_h.squeeze(),
        'df_niveles_ebro': datos.df_pct_ebre_h.squeeze(),
        'df_energia_turbinada_mensual_internes': datos.energia_turbinada_mensual_internes,
        'df_energia_turbinada_mensual_ebre': datos.energia_turbinada_mensual_ebre,
        'df_nivel_si': hydro_base_level,
        
        # Paràmetres físics
        'max_capacity_int': max_capacity_int,
        'max_capacity_ebro': max_capacity_ebro,
        'potencia_max_int': potencia_max_hidraulica_int,
        'potencia_max_ebro': potencia_max_hidraulica_ebro,
        'sensibility_int': sensibility_int,
        'sensibility_ebro': sensibility_ebro,
        
        # Paràmetres hídrics
        'consumo_base_diario_estacional_hm3': consumo_base_diario_estacional,
        'save_hm3_per_mwh': 1/desal_sensibility,
        
        # Precomputats
        'precomputed': precomputed,
        
        # Variables fixes
        'nucleares_activas': [True, True, True],
        'potencia_cogeneracion': 943.503,
        'duracion_horas': 4,
        'potencia_autoconsumo': 1188.6267,
        'demanda_electrica': 1,
        'CF_eolica_obj': None,
        'usar_CF_automatic': True,
        'trend_time_window': 24*30,
        'k_deriv': 0,
        'regen_base_pct': 0.5,
        'llindar_activacio_regen_max': 1,
        'umbrales_sequia': increments_a_llindars(16, 9, 15, 20),  # Valors per defecte
    }
    
    # Ejecutar DOE completo
    doe_resultados = ejecutar_doe_completo(procesar_escenario, datos_base)
    
    
#%%
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def afegir_pestanyes_yates_excel(ruta_excel_input, ruta_excel_output=None):
    """
    Llegeix un Excel de DOE, agafa la pestanya 'Experiments' i afegeix
    una nova pestanya per a cada variable resposta amb les fórmules de Yates.
    """
    
    # Si no s'especifica output, sobreescrivim el mateix fitxer
    if ruta_excel_output is None:
        ruta_excel_output = ruta_excel_input

    print(f"Llegint dades de: {ruta_excel_input} ...")
    
    # 1. Llegim les dades amb Pandas per facilitat (només la pestanya Experiments)
    try:
        df = pd.read_excel(ruta_excel_input, sheet_name='Experiments')
    except ValueError:
        print("Error: No s'ha trobat la pestanya 'Experiments'. Verifica el nom.")
        return

    # Configuració de columnes basada en el teu fitxer
    cols_factors = ['fotovoltaica', 'eolica', 'baterias', 'desalacion', 'min_run_hours', 'midpoint', 'overflow_threshold']
    # Llista de variables resposta (les detectem o les definim manualment segons el teu fitxer)
    cols_resposta = ['min_nivell', 'termica_imports', 'excedents', 'costos', 'factor_capacitat', 'amplitud_estacional', 'msd_nivell']
    
    n_runs = len(df)
    n_factors = 7  # 2^7 = 128
    
    if n_runs != 128:
        print(f"Alerta: Hi ha {n_runs} files. L'algoritme de Yates per 7 factors requereix exactament 128 experiments.")

    # 2. Carreguem el llibre Excel existent amb openpyxl per poder AFEGIR pestanyes
    wb = load_workbook(ruta_excel_input)
    
    print("Generant pestanyes de Yates...")

    # Bucle per a cada variable de resposta
    for respuesta in cols_resposta:
        # Nom de la nova pestanya (Excel limita a 31 caràcters)
        nom_pestanya = f"Yates_{respuesta}"[:31]
        
        # Si la pestanya ja existeix, l'esborrem per refer-la (opcional)
        if nom_pestanya in wb.sheetnames:
            del wb[nom_pestanya]
        
        ws = wb.create_sheet(title=nom_pestanya)
        
        # --- CAPÇALERES ---
        headers = ['Ordre', 'Etiqueta'] + cols_factors + ['Resultat (y)']
        # Columnes d'iteració Yates
        for i in range(1, n_factors + 1):
            headers.append(f'Yates_{i}')
        headers.append('Efecte (Contrast)')
        
        ws.append(headers)
        
        # Determinem índexs de columna per a les fórmules
        # Ordre(A), Etiqueta(B), Factors(C-I), Resultat(J) -> Total 2 + 7 + 1 = 10 columnes inicials
        # La columna 'Resultat (y)' és la columna 10 (lletra J)
        col_resultat_idx = 2 + len(cols_factors) + 1 
        
        # --- ESCRIURE DADES BASE ---
        for index, row in df.iterrows():
            fila_excel = index + 2
            # Construïm la llista de valors per la fila
            valors = [row['experimento'], row['etiqueta']]
            for f in cols_factors:
                valors.append(row[f])
            
            # Afegim el valor de la resposta actual
            valors.append(row[respuesta])
            
            # Escrivim a la cel·la
            for c_idx, valor in enumerate(valors, start=1):
                ws.cell(row=fila_excel, column=c_idx, value=valor)
        
        # --- GENERAR FÓRMULES DE YATES ---
        start_col_yates = col_resultat_idx + 1 # Comença a la columna K (11)
        
        for step in range(n_factors):
            current_col_idx = start_col_yates + step
            prev_col_idx = current_col_idx - 1
            
            col_lletra_prev = get_column_letter(prev_col_idx)
            
            half = n_runs // 2
            
            for i in range(half):
                # Fila destí (superior i inferior)
                row_dest_top = 2 + i 
                row_dest_bot = 2 + i + half
                
                # Files origen (parelles: 1 i 2, 3 i 4...)
                # Recorda: i=0 -> files 2 i 3 de l'Excel
                row_src_1 = 2 + (i * 2)
                row_src_2 = 2 + (i * 2) + 1
                
                # Fórmules segons la teva descripció
                # Top: Suma (1 + 2)
                formula_suma = f"={col_lletra_prev}{row_src_1}+{col_lletra_prev}{row_src_2}"
                # Bot: Resta (2 - 1)
                formula_resta = f"={col_lletra_prev}{row_src_2}-{col_lletra_prev}{row_src_1}"
                
                ws.cell(row=row_dest_top, column=current_col_idx, value=formula_suma)
                ws.cell(row=row_dest_bot, column=current_col_idx, value=formula_resta)

        # --- COLUMNA FINAL: CÀLCUL DELS EFECTES ---
        col_efectes_idx = start_col_yates + n_factors
        col_lletra_last_yates = get_column_letter(col_efectes_idx - 1)
        
        # Fila 2 (Mitjana): Dividir per N (128)
        ws.cell(row=2, column=col_efectes_idx, value=f"={col_lletra_last_yates}2/{n_runs}")
        
        # Resta de files: Dividir per N/2 (64)
        for r in range(3, n_runs + 2):
            ws.cell(row=r, column=col_efectes_idx, value=f"={col_lletra_last_yates}{r}/{n_runs/2}")

    # 3. Guardar el fitxer
    wb.save(ruta_excel_output)
    print(f"Fet! S'ha actualitzat el fitxer: {ruta_excel_output}")

# --- EXEMPLE D'ÚS ---
# Assegura't que el fitxer està a la carpeta o posa la ruta completa
nom_fitxer = 'resultats_DOE.xlsx' 
afegir_pestanyes_yates_excel(nom_fitxer, 'resultats_DOE_Yates.xlsx')

#%%

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import combinations

def afegir_totes_interaccions(ruta_excel, sheet_name="Resum_Interaccions"):
    """
    Calcula i afegeix les interaccions de 2, 3, 4... fins a 7 factors
    a una mateixa pestanya, una sota l'altra.
    """
    print(f"Processant interaccions d'ordre superior per: {ruta_excel} ...")
    
    # 1. Carregar dades
    try:
        df = pd.read_excel(ruta_excel, sheet_name='Experiments')
    except ValueError:
        print("Error: No s'ha trobat la pestanya 'Experiments'.")
        return

    # Configuració
    cols_factors = ['fotovoltaica', 'eolica', 'baterias', 'desalacion', 'min_run_hours', 'midpoint', 'overflow_threshold']
    # Mapetgem els noms reals (etiquetes curtes) per les capçaleres (A, B, C...) per fer-ho més llegible
    lletres_factors = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    mapa_factors = dict(zip(cols_factors, lletres_factors))
    
    cols_resposta = ['min_nivell', 'termica_imports', 'excedents', 'costos', 'factor_capacitat', 'amplitud_estacional', 'msd_nivell']
    
    n_runs = len(df)
    divisor = n_runs / 2  # Per calcular l'Efecte (Contrast / (N/2))
    
    # 2. Preparar l'Excel per escriure
    wb = load_workbook(ruta_excel)
    
    # Si la pestanya ja existeix, decidim si la volem netejar o afegir al final. 
    # Per seguretat, creem una de nova o netegem l'existent.
    if sheet_name in wb.sheetnames:
        # Opcional: Esborrar per regenerar net
        del wb[sheet_name]
    
    ws = wb.create_sheet(title=sheet_name, index=3)
    
    # Índex de fila actual per anar escrivint cap avall
    current_row = 1

    # 3. Bucle per ordre d'interacció (k=2 fins a k=7)
    # k=2 (Parelles), k=3 (Trios), etc.
    for k in range(2, len(cols_factors) + 1):
        
        print(f"Calculant interaccions d'ordre {k}...")
        
        # Títol de la secció
        ws.cell(row=current_row, column=1, value=f"--- INTERACCIONS DE {k} FACTORS ---")
        ws.cell(row=current_row, column=1).style = 'Title' # Estil opcional si tens temes
        current_row += 1
        
        # Generar combinacions (ex: ('fotovoltaica', 'eolica', 'baterias'))
        combs = list(combinations(cols_factors, k))
        
        # Diccionari per guardar resultats d'aquest bloc: { 'NomInteraccio': [valors per cada objectiu] }
        resultats_bloc = {'Objectiu': cols_resposta}
        
        for comb in combs:
            # 1. Construir nom curt (ex: ABC)
            nom_columna = "".join([mapa_factors[f] for f in comb])
            
            # 2. Calcular el signe combinat (multiplicació dels signes dels factors)
            # Això crea una columna temporal que és el producte de les columnes dels factors implicats
            # Si A=1, B=-1, C=1 -> Signe = -1
            signe_combinat = df[list(comb)].prod(axis=1)
            
            valors_interaccio = []
            
            # 3. Calcular valor per cada variable resposta
            for obj in cols_resposta:
                # Fórmula Vectoritzada (Molt ràpida): Suma(y * signe) / Divisor
                # Això és matemàticament idèntic a (Mitjana_positius - Mitjana_negatius)
                suma_contrast = (df[obj] * signe_combinat).sum()
                efecte = suma_contrast / divisor
                valors_interaccio.append(efecte)
            
            # Guardem la columna
            resultats_bloc[nom_columna] = valors_interaccio

        # Convertir a DataFrame per facilitar l'escriptura
        df_bloc = pd.DataFrame(resultats_bloc)
        
        # Escriure el DataFrame a l'Excel a partir de current_row
        for r_idx, row in enumerate(dataframe_to_rows(df_bloc, index=False, header=True), start=current_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Actualitzar el punter de fila per al següent bloc (+2 per deixar espai)
        current_row += len(df_bloc) + 3

    # 4. Guardar
    wb.save(ruta_excel)
    print(f"Fet! Totes les interaccions (2-{len(cols_factors)}) s'han guardat a la pestanya '{sheet_name}'.")

# --- EXECUCIÓ ---
# Canvia el nom del fitxer pel teu
nom_fitxer = 'resultats_DOE_Yates.xlsx' 
afegir_totes_interaccions(nom_fitxer)